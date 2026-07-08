"""
build_price_db.py
Generates docs/price_database_seed.xlsx with 3 sheets:
  Sheet 1: Price Guide  (wide format, merged headers, colour-coded by category)
  Sheet 2: DB Seed      (flat format, 4 rows per item for MY_KL / MY_PG / MY_JB / SG)
  Sheet 3: Summary      (one row per category with stats)

Run from the project root:
    python scripts/build_price_db.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------
ITEMS = [
    # Construction (15)
    {"cat":"Construction","subcat":"Brick Wall","name":"Brick wall clay c/w plaster S&I","method":"Standard Clay Brick","unit":"sqft","supply":"supply_install","kl_min":20,"kl_max":32,"sg_min":9.50,"sg_max":19.00},
    {"cat":"Construction","subcat":"Brick Wall","name":"Brick wall w/o plastering S&I","method":"Concrete Block","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":25,"sg_min":7.00,"sg_max":13.00},
    {"cat":"Construction","subcat":"Brick Wall","name":"Lightweight AAC block wall S&I","method":"Lightweight Block (AAC)","unit":"sqft","supply":"supply_install","kl_min":24,"kl_max":35,"sg_min":None,"sg_max":None},
    {"cat":"Construction","subcat":"RC Floor Slab","name":"RC Floor Slab 150mm G25","method":"150mm G25","unit":"sqft","supply":"supply_install","kl_min":35,"kl_max":55,"sg_min":25.00,"sg_max":45.00},
    {"cat":"Construction","subcat":"RC Floor Slab","name":"RC Floor Slab 150mm G30","method":"150mm G30","unit":"sqft","supply":"supply_install","kl_min":38,"kl_max":60,"sg_min":28.00,"sg_max":50.00},
    {"cat":"Construction","subcat":"Screeding","name":"Concrete flooring c/w BRC","method":"Standard","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":35,"sg_min":8.00,"sg_max":18.00},
    {"cat":"Construction","subcat":"Screeding","name":"Sand-cement screed","method":"Standard Screed","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":15,"sg_min":2.00,"sg_max":5.00},
    {"cat":"Construction","subcat":"Plastering","name":"Standard plastering","method":"Standard Plaster","unit":"sqft","supply":"supply_install","kl_min":5,"kl_max":10,"sg_min":2.50,"sg_max":6.00},
    {"cat":"Construction","subcat":"Kerb","name":"Kerb brick c/w plaster S&I","method":"Brick/Concrete","unit":"ft","supply":"supply_install","kl_min":40,"kl_max":80,"sg_min":80.00,"sg_max":180.00},
    {"cat":"Construction","subcat":"Extension Work","name":"RC slab extension","method":"RC Slab Extension","unit":"sqft","supply":"supply_install","kl_min":40,"kl_max":65,"sg_min":35.00,"sg_max":65.00},
    {"cat":"Construction","subcat":"Structural","name":"Retaining wall","method":"RC Retaining","unit":"sqft","supply":"supply_install","kl_min":55,"kl_max":95,"sg_min":30.00,"sg_max":80.00},
    {"cat":"Construction","subcat":"Structural","name":"RC column/beam","method":"Reinforced Concrete","unit":"ft","supply":"supply_install","kl_min":120,"kl_max":220,"sg_min":350.00,"sg_max":800.00},
    {"cat":"Construction","subcat":"Staircase","name":"RC staircase cast in-situ","method":"Reinforced Concrete","unit":"unit","supply":"supply_install","kl_min":8000,"kl_max":15000,"sg_min":8000,"sg_max":20000},
    {"cat":"Construction","subcat":"Staircase","name":"Mild steel staircase","method":"Mild Steel I-beam","unit":"unit","supply":"supply_install","kl_min":5000,"kl_max":10000,"sg_min":3500,"sg_max":10000},
    {"cat":"Construction","subcat":"Drainage","name":"Concrete drainage channel","method":"U-channel Concrete","unit":"lm","supply":"supply_install","kl_min":40,"kl_max":80,"sg_min":80.00,"sg_max":200.00},
    # Demolition (8)
    {"cat":"Demolition","subcat":"Floor Hacking","name":"Hack floor tiles","method":"Standard Floor","unit":"sqft","supply":"labour_only","kl_min":2.5,"kl_max":4,"sg_min":3.00,"sg_max":6.00},
    {"cat":"Demolition","subcat":"Wall Hacking","name":"Hack wall tiles","method":"Wall Tiles","unit":"sqft","supply":"labour_only","kl_min":3,"kl_max":6,"sg_min":3.00,"sg_max":7.00},
    {"cat":"Demolition","subcat":"Wall Hacking","name":"Hack non-structural wall","method":"Non-structural","unit":"sqft","supply":"labour_only","kl_min":6,"kl_max":9,"sg_min":6.00,"sg_max":18.00},
    {"cat":"Demolition","subcat":"Wall Hacking","name":"Demolish structural wall","method":"Structural","unit":"sqft","supply":"labour_only","kl_min":10,"kl_max":16,"sg_min":15.00,"sg_max":40.00},
    {"cat":"Demolition","subcat":"Ceiling Removal","name":"Remove false ceiling","method":"Full system","unit":"sqft","supply":"labour_only","kl_min":3,"kl_max":5,"sg_min":3.00,"sg_max":6.00},
    {"cat":"Demolition","subcat":"Fixture Removal","name":"Remove door/window frame","method":"Standard Door","unit":"unit","supply":"labour_only","kl_min":120,"kl_max":200,"sg_min":80.00,"sg_max":200.00},
    {"cat":"Demolition","subcat":"Disposal","name":"Debris disposal per lorry","method":"Lorry Load","unit":"load","supply":"labour_only","kl_min":350,"kl_max":550,"sg_min":300.00,"sg_max":800.00},
    {"cat":"Demolition","subcat":"Fixture Removal","name":"Remove sanitary ware","method":"Sanitary Ware","unit":"unit","supply":"labour_only","kl_min":150,"kl_max":300,"sg_min":None,"sg_max":None},
    # Tiling (16)
    {"cat":"Tiling","subcat":"Floor Tile S&I","name":"Floor tiles S&I 300x300","method":"Porcelain 300x300","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":22,"sg_min":9.00,"sg_max":13.00},
    {"cat":"Tiling","subcat":"Floor Tile S&I","name":"Floor tiles S&I 600x600","method":"Porcelain 600x600","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":28,"sg_min":10.00,"sg_max":15.00},
    {"cat":"Tiling","subcat":"Floor Tile S&I","name":"Floor tiles S&I 800x800","method":"Porcelain 800x800","unit":"sqft","supply":"supply_install","kl_min":20,"kl_max":30,"sg_min":13.00,"sg_max":20.00},
    {"cat":"Tiling","subcat":"Floor Tile S&I","name":"Floor tiles S&I 600x1200","method":"Porcelain 1200x600","unit":"sqft","supply":"supply_install","kl_min":25,"kl_max":35,"sg_min":14.00,"sg_max":22.00},
    {"cat":"Tiling","subcat":"Labour Only","name":"Floor tiles labour only 300x600","method":"Standard Format","unit":"sqft","supply":"labour_only","kl_min":6,"kl_max":10,"sg_min":6.00,"sg_max":8.00},
    {"cat":"Tiling","subcat":"Labour Only","name":"Floor tiles labour only 600x1200+","method":"Large Format","unit":"sqft","supply":"labour_only","kl_min":8,"kl_max":14,"sg_min":8.00,"sg_max":11.00},
    {"cat":"Tiling","subcat":"Wall Tile S&I","name":"Wall tiles S&I 300x600","method":"Ceramic/Porcelain 300x600","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":26,"sg_min":10.00,"sg_max":18.00},
    {"cat":"Tiling","subcat":"Wall Tile S&I","name":"Wall tiles S&I large format 600x1200","method":"Ceramic/Porcelain 600x1200","unit":"sqft","supply":"supply_install","kl_min":25,"kl_max":35,"sg_min":14.00,"sg_max":25.00},
    {"cat":"Tiling","subcat":"Specialty Tile","name":"Mosaic tiles S&I","method":"Mosaic/Feature","unit":"sqft","supply":"supply_install","kl_min":25,"kl_max":45,"sg_min":8.00,"sg_max":20.00},
    {"cat":"Tiling","subcat":"Specialty Tile","name":"Outdoor anti-slip tiles S&I","method":"Anti-slip Porcelain","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":30,"sg_min":10.00,"sg_max":22.00},
    {"cat":"Tiling","subcat":"Grouting","name":"Tile grouting standard","method":"Cementitious Grout","unit":"sqft","supply":"labour_only","kl_min":1.5,"kl_max":2.5,"sg_min":1.00,"sg_max":3.00},
    {"cat":"Tiling","subcat":"Grouting","name":"Tile grouting epoxy","method":"Epoxy Grout","unit":"sqft","supply":"labour_only","kl_min":3.5,"kl_max":5.5,"sg_min":5.00,"sg_max":12.00},
    {"cat":"Tiling","subcat":"Specialty Tile","name":"Natural marble tile S&I","method":"Natural Marble","unit":"sqft","supply":"supply_install","kl_min":35,"kl_max":80,"sg_min":None,"sg_max":None},
    {"cat":"Tiling","subcat":"Specialty Tile","name":"Natural granite tile S&I","method":"Natural Granite","unit":"sqft","supply":"supply_install","kl_min":30,"kl_max":65,"sg_min":None,"sg_max":None},
    {"cat":"Tiling","subcat":"Labour Only","name":"Wall tiles labour only","method":"Standard Wall","unit":"sqft","supply":"labour_only","kl_min":7,"kl_max":12,"sg_min":None,"sg_max":None},
    {"cat":"Tiling","subcat":"Labour Only","name":"Mosaic tiles labour only","method":"Mosaic Sheet","unit":"sqft","supply":"labour_only","kl_min":10,"kl_max":16,"sg_min":None,"sg_max":None},
    # Waterproofing (7)
    {"cat":"Waterproofing","subcat":"Bathroom","name":"Bathroom floor waterproofing","method":"Cementitious","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":18,"sg_min":2.00,"sg_max":5.00},
    {"cat":"Waterproofing","subcat":"Bathroom","name":"Bathroom wall waterproofing","method":"Membrane Sheet","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":13,"sg_min":2.00,"sg_max":5.00},
    {"cat":"Waterproofing","subcat":"Roof","name":"Flat roof torch-on membrane","method":"Torch-on Membrane","unit":"sqft","supply":"supply_install","kl_min":12,"kl_max":25,"sg_min":3.00,"sg_max":8.00},
    {"cat":"Waterproofing","subcat":"Balcony","name":"Balcony/patio waterproofing","method":"Cementitious","unit":"sqft","supply":"supply_install","kl_min":10,"kl_max":20,"sg_min":2.50,"sg_max":5.00},
    {"cat":"Waterproofing","subcat":"Tank","name":"Water tank waterproofing","method":"Epoxy Lining","unit":"sqft","supply":"supply_install","kl_min":6,"kl_max":12,"sg_min":3.00,"sg_max":7.00},
    {"cat":"Waterproofing","subcat":"Planter","name":"Planter box waterproofing","method":"Membrane + Drain","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":13,"sg_min":None,"sg_max":None},
    {"cat":"Waterproofing","subcat":"Specialty","name":"Crystalline waterproofing","method":"Crystalline","unit":"sqft","supply":"supply_install","kl_min":12,"kl_max":22,"sg_min":None,"sg_max":None},
    # Electrical (19)
    {"cat":"Electrical","subcat":"Lighting","name":"Lighting point S&I","method":"Standard","unit":"pt","supply":"supply_install","kl_min":55,"kl_max":100,"sg_min":55.00,"sg_max":150.00},
    {"cat":"Electrical","subcat":"Lighting","name":"Lighting point high ceiling S&I","method":"High Ceiling","unit":"pt","supply":"supply_install","kl_min":80,"kl_max":130,"sg_min":80.00,"sg_max":200.00},
    {"cat":"Electrical","subcat":"Power Points","name":"13A socket point S&I","method":"Standard 13A","unit":"pt","supply":"supply_install","kl_min":85,"kl_max":150,"sg_min":55.00,"sg_max":130.00},
    {"cat":"Electrical","subcat":"Power Points","name":"15A socket point S&I","method":"Heavy-duty 15A","unit":"pt","supply":"supply_install","kl_min":100,"kl_max":170,"sg_min":100.00,"sg_max":180.00},
    {"cat":"Electrical","subcat":"Power Points","name":"20A socket point S&I","method":"20A Heavy Duty","unit":"pt","supply":"supply_install","kl_min":120,"kl_max":200,"sg_min":90.00,"sg_max":160.00},
    {"cat":"Electrical","subcat":"Power Points","name":"USB socket point S&I","method":"USB-A/C Combo","unit":"pt","supply":"supply_install","kl_min":120,"kl_max":200,"sg_min":60.00,"sg_max":130.00},
    {"cat":"Electrical","subcat":"Fan Points","name":"Ceiling fan point S&I","method":"Standard Fan Pt","unit":"pt","supply":"supply_install","kl_min":100,"kl_max":200,"sg_min":55.00,"sg_max":160.00},
    {"cat":"Electrical","subcat":"Fan Points","name":"Exhaust fan point S&I","method":"Wall/Ceiling Extract","unit":"pt","supply":"supply_install","kl_min":80,"kl_max":150,"sg_min":40.00,"sg_max":100.00},
    {"cat":"Electrical","subcat":"Lighting","name":"Cove light point S&I","method":"LED Strip + Profile","unit":"pt","supply":"supply_install","kl_min":120,"kl_max":250,"sg_min":30.00,"sg_max":80.00},
    {"cat":"Electrical","subcat":"Lighting","name":"Downlight cutout + wiring","method":"LED Downlight","unit":"pt","supply":"supply_install","kl_min":35,"kl_max":60,"sg_min":30.00,"sg_max":100.00},
    {"cat":"Electrical","subcat":"Re-locate/Rewire","name":"Re-locate existing point","method":"Standard Reloc","unit":"pt","supply":"supply_install","kl_min":150,"kl_max":300,"sg_min":25.00,"sg_max":80.00},
    {"cat":"Electrical","subcat":"Data Points","name":"TV/data point S&I","method":"Coaxial + Face Plate","unit":"pt","supply":"supply_install","kl_min":150,"kl_max":300,"sg_min":80.00,"sg_max":130.00},
    {"cat":"Electrical","subcat":"Distribution Board","name":"DB box 12-way S&I","method":"12-way MCB","unit":"unit","supply":"supply_install","kl_min":900,"kl_max":1400,"sg_min":280.00,"sg_max":450.00},
    {"cat":"Electrical","subcat":"Distribution Board","name":"DB box 18-way S&I","method":"18-way MCB","unit":"unit","supply":"supply_install","kl_min":800,"kl_max":1500,"sg_min":300.00,"sg_max":600.00},
    {"cat":"Electrical","subcat":"Distribution Board","name":"DB box 36-way S&I","method":"36-way MCB","unit":"unit","supply":"supply_install","kl_min":1200,"kl_max":2200,"sg_min":600.00,"sg_max":1000.00},
    {"cat":"Electrical","subcat":"Distribution Board","name":"DB box 48-way S&I","method":"48-way MCB","unit":"unit","supply":"supply_install","kl_min":2500,"kl_max":3500,"sg_min":800.00,"sg_max":1400.00},
    {"cat":"Electrical","subcat":"Re-locate/Rewire","name":"Rewiring per room","method":"Complete Rewire","unit":"room","supply":"labour_only","kl_min":3500,"kl_max":8000,"sg_min":800.00,"sg_max":2500.00},
    {"cat":"Electrical","subcat":"Smart Home","name":"Digital door lock S&I","method":"Digital Lock","unit":"unit","supply":"supply_install","kl_min":500,"kl_max":1500,"sg_min":None,"sg_max":None},
    {"cat":"Electrical","subcat":"Smart Home","name":"Smart switch S&I","method":"Smart Switch","unit":"pt","supply":"supply_install","kl_min":120,"kl_max":250,"sg_min":None,"sg_max":None},
    # Plumbing (13)
    {"cat":"Plumbing","subcat":"Piping","name":"Piping PPR/copper per point","method":"PPR 20-32mm","unit":"pt","supply":"supply_install","kl_min":70,"kl_max":110,"sg_min":200.00,"sg_max":500.00},
    {"cat":"Plumbing","subcat":"Piping","name":"Replace piping per bathroom","method":"Chase + Pipe","unit":"room","supply":"supply_install","kl_min":1500,"kl_max":5000,"sg_min":1500.00,"sg_max":3000.00},
    {"cat":"Plumbing","subcat":"Sanitary Ware","name":"Basin mixer tap S&I","method":"Standard Basin","unit":"unit","supply":"supply_install","kl_min":300,"kl_max":600,"sg_min":100.00,"sg_max":250.00},
    {"cat":"Plumbing","subcat":"Kitchen","name":"Kitchen sink + tap S&I","method":"SS Single Bowl","unit":"unit","supply":"supply_install","kl_min":350,"kl_max":650,"sg_min":300.00,"sg_max":700.00},
    {"cat":"Plumbing","subcat":"Sanitary Ware","name":"WC floor-mount S&I","method":"Floor Mount","unit":"unit","supply":"supply_install","kl_min":400,"kl_max":800,"sg_min":420.00,"sg_max":900.00},
    {"cat":"Plumbing","subcat":"Sanitary Ware","name":"Wall-hung WC S&I","method":"Wall Hung","unit":"unit","supply":"supply_install","kl_min":800,"kl_max":1500,"sg_min":600.00,"sg_max":1200.00},
    {"cat":"Plumbing","subcat":"Shower/Bath","name":"Rain shower set S&I","method":"Overhead + Handheld","unit":"unit","supply":"supply_install","kl_min":400,"kl_max":900,"sg_min":300.00,"sg_max":800.00},
    {"cat":"Plumbing","subcat":"Water Heater","name":"Instant water heater S&I","method":"Instant Heater","unit":"unit","supply":"supply_install","kl_min":200,"kl_max":500,"sg_min":280.00,"sg_max":500.00},
    {"cat":"Plumbing","subcat":"Water Heater","name":"Storage water heater S&I","method":"Storage 25-40L","unit":"unit","supply":"supply_install","kl_min":800,"kl_max":1500,"sg_min":400.00,"sg_max":700.00},
    {"cat":"Plumbing","subcat":"Drainage","name":"Floor trap S&I","method":"SS/Brass Floor Trap","unit":"unit","supply":"supply_install","kl_min":90,"kl_max":150,"sg_min":150.00,"sg_max":350.00},
    {"cat":"Plumbing","subcat":"Drainage","name":"SS linear channel drain","method":"SS Linear Channel","unit":"unit","supply":"supply_install","kl_min":200,"kl_max":450,"sg_min":None,"sg_max":None},
    {"cat":"Plumbing","subcat":"Kitchen","name":"Water filter under-sink","method":"Under-sink Filter","unit":"unit","supply":"supply_install","kl_min":400,"kl_max":900,"sg_min":None,"sg_max":None},
    {"cat":"Plumbing","subcat":"Shower/Bath","name":"Shower screen tempered S&I","method":"Tempered Glass","unit":"unit","supply":"supply_install","kl_min":600,"kl_max":1200,"sg_min":None,"sg_max":None},
    # Painting (10)
    {"cat":"Painting","subcat":"Interior","name":"Interior wall 2-coat emulsion S&I","method":"Emulsion Paint","unit":"sqft","supply":"supply_install","kl_min":2.5,"kl_max":5,"sg_min":2.00,"sg_max":5.00},
    {"cat":"Painting","subcat":"Interior","name":"Interior wall 3-coat emulsion S&I","method":"Premium Emulsion","unit":"sqft","supply":"supply_install","kl_min":3,"kl_max":6,"sg_min":2.50,"sg_max":6.00},
    {"cat":"Painting","subcat":"Interior","name":"Ceiling paint 2-coat S&I","method":"Ceiling Paint","unit":"sqft","supply":"supply_install","kl_min":1.3,"kl_max":2,"sg_min":1.50,"sg_max":3.00},
    {"cat":"Painting","subcat":"Specialty","name":"Feature wall texture paint","method":"Textured Finish","unit":"sqft","supply":"supply_install","kl_min":4.5,"kl_max":7.5,"sg_min":3.00,"sg_max":8.00},
    {"cat":"Painting","subcat":"Exterior","name":"Exterior painting weather shield","method":"Weather Paint","unit":"sqft","supply":"supply_install","kl_min":3,"kl_max":6,"sg_min":2.50,"sg_max":6.00},
    {"cat":"Painting","subcat":"Interior","name":"Skim coat + paint","method":"Skim + Paint","unit":"sqft","supply":"supply_install","kl_min":2.5,"kl_max":4,"sg_min":2.50,"sg_max":5.00},
    {"cat":"Painting","subcat":"Interior","name":"Re-paint existing walls","method":"Spot Repair","unit":"sqft","supply":"supply_install","kl_min":1.8,"kl_max":4,"sg_min":1.00,"sg_max":2.00},
    {"cat":"Painting","subcat":"Interior","name":"Touch-up painting","method":"Colour Matching","unit":"lump sum","supply":"supply_install","kl_min":300,"kl_max":1500,"sg_min":200.00,"sg_max":800.00},
    {"cat":"Painting","subcat":"Specialty","name":"Clear lacquer / spray","method":"Clear Lacquer","unit":"sqft","supply":"supply_install","kl_min":3.5,"kl_max":6,"sg_min":None,"sg_max":None},
    {"cat":"Painting","subcat":"Exterior","name":"Anti-fungal exterior coat","method":"Anti-fungal","unit":"sqft","supply":"supply_install","kl_min":3.5,"kl_max":6,"sg_min":None,"sg_max":None},
    # False Ceiling (10)
    {"cat":"False Ceiling","subcat":"Plasterboard","name":"Flat plasterboard ceiling S&I","method":"9mm Plasterboard","unit":"sqft","supply":"supply_install","kl_min":10,"kl_max":18,"sg_min":3.75,"sg_max":6.00},
    {"cat":"False Ceiling","subcat":"Plasterboard","name":"L-box cove ceiling S&I","method":"L-box/Tray Design","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":35,"sg_min":32.00,"sg_max":45.00},
    {"cat":"False Ceiling","subcat":"Plasterboard","name":"L-box cove with LED wiring","method":"Plasterboard + LED","unit":"sqft","supply":"supply_install","kl_min":14,"kl_max":22,"sg_min":40.00,"sg_max":55.00},
    {"cat":"False Ceiling","subcat":"Specialty Ceiling","name":"Calcium silicate ceiling S&I","method":"Calcium Silicate","unit":"sqft","supply":"supply_install","kl_min":12,"kl_max":20,"sg_min":4.00,"sg_max":7.00},
    {"cat":"False Ceiling","subcat":"Plasterboard","name":"Plaster of Paris (POP) cornice","method":"Plaster Cornice","unit":"lm","supply":"supply_install","kl_min":8,"kl_max":14,"sg_min":20.00,"sg_max":40.00},
    {"cat":"False Ceiling","subcat":"Plasterboard","name":"Polyurethane (PU) cornice S&I","method":"Polyurethane","unit":"lm","supply":"supply_install","kl_min":12,"kl_max":20,"sg_min":None,"sg_max":None},
    {"cat":"False Ceiling","subcat":"Partition","name":"Partition wall single layer gypsum","method":"Plasterboard 1-side","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":30,"sg_min":4.20,"sg_max":6.00},
    {"cat":"False Ceiling","subcat":"Partition","name":"Partition wall double layer gypsum","method":"Plasterboard 2-side","unit":"sqft","supply":"supply_install","kl_min":28,"kl_max":45,"sg_min":5.50,"sg_max":8.25},
    {"cat":"False Ceiling","subcat":"Specialty Ceiling","name":"PVC ceiling panel S&I","method":"Mineral Fibre Tile","unit":"sqft","supply":"supply_install","kl_min":6,"kl_max":12,"sg_min":3.50,"sg_max":6.50},
    {"cat":"False Ceiling","subcat":"Partition","name":"Glass + plasterboard partition","method":"Plasterboard + Glass","unit":"sqft","supply":"supply_install","kl_min":35,"kl_max":60,"sg_min":None,"sg_max":None},
    # Carpentry (27)
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen cabinet upper wall-hung laminate","method":"Laminate (wall cabinet)","unit":"ft","supply":"supply_install","kl_min":300,"kl_max":650,"sg_min":100.00,"sg_max":180.00},
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen cabinet lower base laminate","method":"Melamine/Laminate","unit":"ft","supply":"supply_install","kl_min":450,"kl_max":950,"sg_min":100.00,"sg_max":180.00},
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen cabinet full height laminate","method":"Laminate (tall unit)","unit":"ft","supply":"supply_install","kl_min":500,"kl_max":1000,"sg_min":200.00,"sg_max":350.00},
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen cabinet solid plywood laminate","method":"Plywood + Blum hardware","unit":"ft","supply":"supply_install","kl_min":500,"kl_max":1000,"sg_min":150.00,"sg_max":260.00},
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen cabinet aluminium frame","method":"ACP/Aluminium frame","unit":"ft","supply":"supply_install","kl_min":400,"kl_max":900,"sg_min":180.00,"sg_max":320.00},
    {"cat":"Carpentry","subcat":"Wardrobe","name":"Wardrobe swing door melamine laminate","method":"Melamine Swing","unit":"ft","supply":"supply_install","kl_min":700,"kl_max":1200,"sg_min":220.00,"sg_max":330.00},
    {"cat":"Carpentry","subcat":"Wardrobe","name":"Wardrobe sliding door melamine laminate","method":"Melamine Sliding","unit":"ft","supply":"supply_install","kl_min":800,"kl_max":1400,"sg_min":240.00,"sg_max":360.00},
    {"cat":"Carpentry","subcat":"Wardrobe","name":"Wardrobe walk-in solid plywood","method":"Plywood + Veneer","unit":"ft","supply":"supply_install","kl_min":900,"kl_max":1600,"sg_min":250.00,"sg_max":380.00},
    {"cat":"Carpentry","subcat":"Display/Feature","name":"TV feature wall","method":"Fluted WPC/MDF panel","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":35,"sg_min":15.00,"sg_max":35.00},
    {"cat":"Carpentry","subcat":"Storage","name":"Shoe cabinet laminate","method":"Standard Shoe","unit":"ft","supply":"supply_install","kl_min":250,"kl_max":550,"sg_min":150.00,"sg_max":300.00},
    {"cat":"Carpentry","subcat":"Vanity","name":"Bathroom vanity cabinet","method":"Standard Vanity","unit":"ft","supply":"supply_install","kl_min":350,"kl_max":700,"sg_min":120.00,"sg_max":200.00},
    {"cat":"Carpentry","subcat":"Display/Feature","name":"Bookshelf / display unit","method":"Wall mount shelving","unit":"ft","supply":"supply_install","kl_min":60,"kl_max":120,"sg_min":150.00,"sg_max":380.00},
    {"cat":"Carpentry","subcat":"Study","name":"Study desk built-in","method":"Built-in laminate","unit":"ft","supply":"supply_install","kl_min":450,"kl_max":750,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Doors","name":"Solid core bedroom door S&I","method":"Solid Core","unit":"unit","supply":"supply_install","kl_min":550,"kl_max":900,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Doors","name":"Hollow core door S&I","method":"Hollow Core","unit":"unit","supply":"supply_install","kl_min":350,"kl_max":550,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Doors","name":"Laminate door S&I","method":"Laminate Finish","unit":"unit","supply":"supply_install","kl_min":400,"kl_max":700,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Doors","name":"Sliding barn door S&I","method":"Sliding Barn","unit":"unit","supply":"supply_install","kl_min":800,"kl_max":1400,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Doors","name":"Pocket door S&I","method":"Pocket/Surface","unit":"unit","supply":"supply_install","kl_min":700,"kl_max":1200,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Doors","name":"Bi-fold door S&I","method":"Folding","unit":"unit","supply":"supply_install","kl_min":800,"kl_max":1500,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen cabinet quartz countertop","method":"15-20mm Quartz top","unit":"ft","supply":"supply_install","kl_min":80,"kl_max":180,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Kitchen Cabinet","name":"Kitchen backsplash tile S&I","method":"Tile Backsplash","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":30,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Wardrobe","name":"Wardrobe mirror panel door","method":"Mirror + Laminate","unit":"ft","supply":"supply_install","kl_min":900,"kl_max":1500,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Storage","name":"Full height cabinet S&I","method":"Full Height Laminate","unit":"ft","supply":"supply_install","kl_min":450,"kl_max":700,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Storage","name":"Bench seat with storage","method":"Bench Box","unit":"ft","supply":"supply_install","kl_min":200,"kl_max":400,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Custom","name":"Built-in bed headboard","method":"Plywood + Cushion","unit":"ft","supply":"supply_install","kl_min":200,"kl_max":450,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Custom","name":"Murphy/fold-down bed","method":"Fold-down Bed","unit":"unit","supply":"supply_install","kl_min":2000,"kl_max":5000,"sg_min":None,"sg_max":None},
    {"cat":"Carpentry","subcat":"Door Frame","name":"Timber door frame S&I","method":"Solid Timber Frame","unit":"unit","supply":"supply_install","kl_min":180,"kl_max":300,"sg_min":None,"sg_max":None},
    # Table Top (4)
    {"cat":"Table Top","subcat":"Stone Top","name":"Quartz countertop S&I","method":"Quartz/Solid Surface","unit":"ft","supply":"supply_install","kl_min":150,"kl_max":350,"sg_min":80.00,"sg_max":150.00},
    {"cat":"Table Top","subcat":"Stone Top","name":"Marble/granite countertop S&I","method":"Natural Granite","unit":"ft","supply":"supply_install","kl_min":200,"kl_max":500,"sg_min":90.00,"sg_max":200.00},
    {"cat":"Table Top","subcat":"Stone Top","name":"Sintered stone countertop S&I","method":"Dekton/Neolith","unit":"ft","supply":"supply_install","kl_min":250,"kl_max":450,"sg_min":None,"sg_max":None},
    {"cat":"Table Top","subcat":"Stone Top","name":"Solid surface countertop S&I","method":"Corian/Similar","unit":"ft","supply":"supply_install","kl_min":280,"kl_max":620,"sg_min":60.00,"sg_max":130.00},
    # Roofing (12)
    {"cat":"Roofing","subcat":"Metal Roof","name":"Metal deck PU foam roofing S&I","method":"Metal Deck + PU","unit":"sqft","supply":"supply_install","kl_min":22,"kl_max":34,"sg_min":5.00,"sg_max":12.00},
    {"cat":"Roofing","subcat":"Polycarbonate","name":"Polycarbonate twinwall roofing S&I","method":"Twinwall PC","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":30,"sg_min":8.00,"sg_max":20.00},
    {"cat":"Roofing","subcat":"Glass Roof","name":"Glass roofing tempered S&I","method":"Tempered Glass","unit":"sqft","supply":"supply_install","kl_min":55,"kl_max":90,"sg_min":18.00,"sg_max":45.00},
    {"cat":"Roofing","subcat":"Tile Roof","name":"Concrete roof tile S&I","method":"Concrete Tile","unit":"sqft","supply":"supply_install","kl_min":6,"kl_max":10,"sg_min":4.00,"sg_max":10.00},
    {"cat":"Roofing","subcat":"Metal Roof","name":"Aluminium composite panel roofing","method":"ACP Roof","unit":"sqft","supply":"supply_install","kl_min":25,"kl_max":40,"sg_min":8.00,"sg_max":18.00},
    {"cat":"Roofing","subcat":"Gutter","name":"Gutter aluminium S&I","method":"Aluminium","unit":"lm","supply":"supply_install","kl_min":35,"kl_max":55,"sg_min":30.00,"sg_max":80.00},
    {"cat":"Roofing","subcat":"Gutter","name":"Downpipe PVC S&I","method":"PVC Half-round","unit":"lm","supply":"supply_install","kl_min":25,"kl_max":42,"sg_min":25.00,"sg_max":60.00},
    {"cat":"Roofing","subcat":"Awning","name":"Metal awning S&I","method":"Metal + Frame","unit":"sqft","supply":"supply_install","kl_min":30,"kl_max":50,"sg_min":10.00,"sg_max":25.00},
    {"cat":"Roofing","subcat":"Awning","name":"Polycarbonate awning S&I","method":"PC + Steel Frame","unit":"sqft","supply":"supply_install","kl_min":25,"kl_max":42,"sg_min":None,"sg_max":None},
    {"cat":"Roofing","subcat":"Insulation","name":"Roof insulation foil bubble","method":"Foil Bubble","unit":"sqft","supply":"supply_install","kl_min":3,"kl_max":6,"sg_min":None,"sg_max":None},
    {"cat":"Roofing","subcat":"Repair","name":"Roof sealant / patch repair","method":"Sealant/Patch","unit":"lump sum","supply":"supply_install","kl_min":300,"kl_max":800,"sg_min":None,"sg_max":None},
    {"cat":"Roofing","subcat":"Tile Roof","name":"Clay roof tile S&I","method":"Clay Tile","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":14,"sg_min":None,"sg_max":None},
    # Aluminium (8)
    {"cat":"Aluminium","subcat":"Window","name":"Casement window aluminium S&I","method":"Standard Casement","unit":"sqft","supply":"supply_install","kl_min":100,"kl_max":250,"sg_min":450.00,"sg_max":1000.00},
    {"cat":"Aluminium","subcat":"Window","name":"Sliding window aluminium S&I","method":"Standard Sliding","unit":"sqft","supply":"supply_install","kl_min":30,"kl_max":48,"sg_min":400.00,"sg_max":900.00},
    {"cat":"Aluminium","subcat":"Door","name":"Sliding door aluminium S&I","method":"Slim Profile","unit":"sqft","supply":"supply_install","kl_min":60,"kl_max":95,"sg_min":800.00,"sg_max":2500.00},
    {"cat":"Aluminium","subcat":"Door","name":"Bi-fold door aluminium S&I","method":"Bi-fold System","unit":"sqft","supply":"supply_install","kl_min":55,"kl_max":80,"sg_min":800.00,"sg_max":2000.00},
    {"cat":"Aluminium","subcat":"Window","name":"Louvre window S&I","method":"Louvre/Jalousie","unit":"sqft","supply":"supply_install","kl_min":28,"kl_max":45,"sg_min":300.00,"sg_max":700.00},
    {"cat":"Aluminium","subcat":"Door","name":"Swing door aluminium S&I","method":"French/Swing","unit":"sqft","supply":"supply_install","kl_min":45,"kl_max":75,"sg_min":380.00,"sg_max":900.00},
    {"cat":"Aluminium","subcat":"Door","name":"Aluminium security grille door S&I","method":"Commercial Front","unit":"sqft","supply":"supply_install","kl_min":50,"kl_max":90,"sg_min":350.00,"sg_max":800.00},
    {"cat":"Aluminium","subcat":"Window","name":"Acoustic double glazed window","method":"Acoustic Double","unit":"sqft","supply":"supply_install","kl_min":55,"kl_max":85,"sg_min":None,"sg_max":None},
    # Glass (9)
    {"cat":"Glass","subcat":"Shower Screen","name":"Shower screen tempered 8mm S&I","method":"Tempered + Fitting","unit":"sqft","supply":"supply_install","kl_min":60,"kl_max":110,"sg_min":450.00,"sg_max":900.00},
    {"cat":"Glass","subcat":"Shower Screen","name":"Shower screen tempered 10mm S&I","method":"Tempered 10mm","unit":"sqft","supply":"supply_install","kl_min":45,"kl_max":90,"sg_min":650.00,"sg_max":1500.00},
    {"cat":"Glass","subcat":"Glass Panel","name":"Fixed glass panel clear tempered","method":"Fixed Panel","unit":"sqft","supply":"supply_install","kl_min":55,"kl_max":100,"sg_min":7.00,"sg_max":18.00},
    {"cat":"Glass","subcat":"Mirror","name":"Mirror S&I","method":"4-5mm Mirror","unit":"sqft","supply":"supply_install","kl_min":22,"kl_max":40,"sg_min":5.00,"sg_max":15.00},
    {"cat":"Glass","subcat":"Partition","name":"Frosted glass partition S&I","method":"Frosted/Sandblast","unit":"sqft","supply":"supply_install","kl_min":35,"kl_max":55,"sg_min":9.00,"sg_max":25.00},
    {"cat":"Glass","subcat":"Door","name":"Frameless tempered glass door S&I","method":"Tempered + Motor","unit":"unit","supply":"supply_install","kl_min":3000,"kl_max":6000,"sg_min":800.00,"sg_max":2500.00},
    {"cat":"Glass","subcat":"Balustrade","name":"Glass balustrade/railing S&I","method":"Tempered + SS","unit":"ft","supply":"supply_install","kl_min":120,"kl_max":250,"sg_min":350.00,"sg_max":1200.00},
    {"cat":"Glass","subcat":"Glass Panel","name":"Reeded/fluted glass panel S&I","method":"Reeded/Fluted","unit":"sqft","supply":"supply_install","kl_min":40,"kl_max":65,"sg_min":None,"sg_max":None},
    {"cat":"Glass","subcat":"Glass Panel","name":"Lacobel / painted tempered glass","method":"Painted Tempered","unit":"sqft","supply":"supply_install","kl_min":30,"kl_max":50,"sg_min":None,"sg_max":None},
    # Flooring (10)
    {"cat":"Flooring","subcat":"Vinyl/SPC","name":"SPC/LVT click-lock flooring S&I","method":"SPC 4-5mm","unit":"sqft","supply":"supply_install","kl_min":6,"kl_max":12,"sg_min":6.00,"sg_max":10.00},
    {"cat":"Flooring","subcat":"Timber","name":"Engineered timber flooring S&I","method":"Engineered","unit":"sqft","supply":"supply_install","kl_min":12,"kl_max":28,"sg_min":8.00,"sg_max":20.00},
    {"cat":"Flooring","subcat":"Timber","name":"Laminate flooring S&I","method":"AC3/AC4","unit":"sqft","supply":"supply_install","kl_min":4.5,"kl_max":7.5,"sg_min":5.00,"sg_max":8.00},
    {"cat":"Flooring","subcat":"Timber","name":"Solid parquet flooring S&I","method":"Solid Parquet","unit":"sqft","supply":"supply_install","kl_min":9,"kl_max":15,"sg_min":10.00,"sg_max":30.00},
    {"cat":"Flooring","subcat":"Vinyl/SPC","name":"Sheet vinyl flooring S&I","method":"Sheet Vinyl","unit":"sqft","supply":"supply_install","kl_min":4,"kl_max":8,"sg_min":2.00,"sg_max":5.00},
    {"cat":"Flooring","subcat":"Vinyl/SPC","name":"Carpet tile S&I","method":"LVT Heavy Duty","unit":"sqft","supply":"supply_install","kl_min":7,"kl_max":12,"sg_min":3.00,"sg_max":8.00},
    {"cat":"Flooring","subcat":"Outdoor","name":"Composite WPC decking S&I","method":"WPC/Acacia","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":18,"sg_min":8.00,"sg_max":14.00},
    {"cat":"Flooring","subcat":"Skirting","name":"Timber/wood skirting S&I","method":"Solid Timber","unit":"lm","supply":"supply_install","kl_min":15,"kl_max":25,"sg_min":5.00,"sg_max":20.00},
    {"cat":"Flooring","subcat":"Timber","name":"Solid hardwood flooring S&I","method":"Chengal/Balau","unit":"sqft","supply":"supply_install","kl_min":25,"kl_max":42,"sg_min":10.00,"sg_max":30.00},
    {"cat":"Flooring","subcat":"Vinyl/SPC","name":"Subfloor levelling compound","method":"Self-levelling","unit":"sqft","supply":"supply_install","kl_min":3.5,"kl_max":10,"sg_min":1.50,"sg_max":5.00},
    # Air Conditioning (13)
    {"cat":"Air Conditioning","subcat":"Wall Split","name":"1.0HP wall-mounted split inverter S&I","method":"1.0HP Inverter","unit":"unit","supply":"supply_install","kl_min":1450,"kl_max":2200,"sg_min":950.00,"sg_max":1300.00},
    {"cat":"Air Conditioning","subcat":"Wall Split","name":"1.5HP wall-mounted split inverter S&I","method":"1.5HP Inverter","unit":"unit","supply":"supply_install","kl_min":1800,"kl_max":3500,"sg_min":1050.00,"sg_max":1700.00},
    {"cat":"Air Conditioning","subcat":"Wall Split","name":"2.0HP wall-mounted split inverter S&I","method":"2.0HP Inverter","unit":"unit","supply":"supply_install","kl_min":2300,"kl_max":3500,"sg_min":1200.00,"sg_max":2200.00},
    {"cat":"Air Conditioning","subcat":"Wall Split","name":"2.5HP wall-mounted split inverter S&I","method":"2.5HP Inverter","unit":"unit","supply":"supply_install","kl_min":2800,"kl_max":4200,"sg_min":1400.00,"sg_max":2500.00},
    {"cat":"Air Conditioning","subcat":"Multi-Split System","name":"System 2 (1 outdoor + 2 indoor) S&I","method":"System 2","unit":"system","supply":"supply_install","kl_min":3500,"kl_max":6100,"sg_min":1960.00,"sg_max":2500.00},
    {"cat":"Air Conditioning","subcat":"Multi-Split System","name":"System 3 (1 outdoor + 3 indoor) S&I","method":"System 3","unit":"system","supply":"supply_install","kl_min":5500,"kl_max":9000,"sg_min":2480.00,"sg_max":5500.00},
    {"cat":"Air Conditioning","subcat":"Multi-Split System","name":"System 4 (1 outdoor + 4 indoor) S&I","method":"System 4","unit":"system","supply":"supply_install","kl_min":7500,"kl_max":13000,"sg_min":2960.00,"sg_max":5000.00},
    {"cat":"Air Conditioning","subcat":"Ceiling Type","name":"Ceiling cassette unit 3.0HP S&I","method":"4-way Cassette 3.0HP","unit":"unit","supply":"supply_install","kl_min":4500,"kl_max":7000,"sg_min":1800.00,"sg_max":4000.00},
    {"cat":"Air Conditioning","subcat":"Ceiling Type","name":"Ducted AC system per room","method":"Ducted System","unit":"room","supply":"supply_install","kl_min":5000,"kl_max":9000,"sg_min":2000.00,"sg_max":4000.00},
    {"cat":"Air Conditioning","subcat":"Maintenance","name":"Aircon basic servicing","method":"Basic Service","unit":"unit","supply":"supply_install","kl_min":80,"kl_max":150,"sg_min":25.00,"sg_max":60.00},
    {"cat":"Air Conditioning","subcat":"Maintenance","name":"Aircon chemical wash","method":"Chemical Wash","unit":"unit","supply":"supply_install","kl_min":150,"kl_max":350,"sg_min":80.00,"sg_max":150.00},
    {"cat":"Air Conditioning","subcat":"Maintenance","name":"Aircon chemical overhaul","method":"Chemical Overhaul","unit":"unit","supply":"supply_install","kl_min":199,"kl_max":299,"sg_min":150.00,"sg_max":450.00},
    {"cat":"Air Conditioning","subcat":"Accessories","name":"AC trunking per room","method":"PVC Trunking","unit":"room","supply":"supply_install","kl_min":200,"kl_max":500,"sg_min":100.00,"sg_max":300.00},
    # Metal Work (10)
    {"cat":"Metal Work","subcat":"Gate","name":"Mild steel gate powder coated S&I","method":"MS Powder Coat","unit":"unit","supply":"supply_install","kl_min":1200,"kl_max":2200,"sg_min":400.00,"sg_max":1000.00},
    {"cat":"Metal Work","subcat":"Gate","name":"Auto gate arm/sliding motor S&I","method":"Arm/Sliding Motor","unit":"unit","supply":"supply_install","kl_min":2500,"kl_max":4500,"sg_min":None,"sg_max":None},
    {"cat":"Metal Work","subcat":"Railing","name":"Mild steel railing S&I","method":"MS Powder Coat","unit":"lm","supply":"supply_install","kl_min":150,"kl_max":260,"sg_min":80.00,"sg_max":200.00},
    {"cat":"Metal Work","subcat":"Railing","name":"Stainless steel railing S&I","method":"Stainless Steel","unit":"lm","supply":"supply_install","kl_min":200,"kl_max":350,"sg_min":150.00,"sg_max":450.00},
    {"cat":"Metal Work","subcat":"Grille/Fence","name":"Metal grille window S&I","method":"MS/Wrought Iron","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":30,"sg_min":8.00,"sg_max":18.00},
    {"cat":"Metal Work","subcat":"Gate","name":"Metal letterbox/pillar S&I","method":"Galvanised","unit":"unit","supply":"supply_install","kl_min":400,"kl_max":800,"sg_min":80.00,"sg_max":250.00},
    {"cat":"Metal Work","subcat":"Railing","name":"Glass + steel railing S&I","method":"Glass + Steel","unit":"lm","supply":"supply_install","kl_min":250,"kl_max":450,"sg_min":None,"sg_max":None},
    {"cat":"Metal Work","subcat":"Grille/Fence","name":"BRC mesh fencing S&I","method":"Galvanised BRC","unit":"lm","supply":"supply_install","kl_min":55,"kl_max":90,"sg_min":40.00,"sg_max":120.00},
    {"cat":"Metal Work","subcat":"Gate","name":"Metal pedestrian gate","method":"MS Powder Coat (pedestrian)","unit":"unit","supply":"supply_install","kl_min":400,"kl_max":800,"sg_min":None,"sg_max":None},
    {"cat":"Metal Work","subcat":"Railing","name":"Wrought iron decorative railing","method":"Decorative Iron","unit":"lm","supply":"supply_install","kl_min":180,"kl_max":320,"sg_min":None,"sg_max":None},
    # Landscape (11)
    {"cat":"Landscape","subcat":"Turfing","name":"Artificial turf S&I","method":"40mm Synthetic","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":35,"sg_min":3.00,"sg_max":12.00},
    {"cat":"Landscape","subcat":"Decking","name":"Composite WPC deck outdoor S&I","method":"WPC Board","unit":"sqft","supply":"supply_install","kl_min":18,"kl_max":30,"sg_min":8.00,"sg_max":16.00},
    {"cat":"Landscape","subcat":"Paving","name":"Concrete/paving stone S&I","method":"Concrete Paver","unit":"sqft","supply":"supply_install","kl_min":8,"kl_max":14,"sg_min":8.00,"sg_max":25.00},
    {"cat":"Landscape","subcat":"Planting","name":"Garden soil + planting","method":"Soil + Shrubs","unit":"sqft","supply":"supply_install","kl_min":5,"kl_max":15,"sg_min":3.00,"sg_max":10.00},
    {"cat":"Landscape","subcat":"Structure","name":"Pergola timber S&I","method":"Treated Timber Frame","unit":"sqft","supply":"supply_install","kl_min":45,"kl_max":90,"sg_min":20.00,"sg_max":50.00},
    {"cat":"Landscape","subcat":"Planting","name":"Irrigation system S&I","method":"Sprinkler + Timer","unit":"lump sum","supply":"supply_install","kl_min":1500,"kl_max":4000,"sg_min":800.00,"sg_max":3500.00},
    {"cat":"Landscape","subcat":"Structure","name":"Feature wall/vertical garden","method":"Modular System","unit":"sqft","supply":"supply_install","kl_min":30,"kl_max":65,"sg_min":15.00,"sg_max":60.00},
    {"cat":"Landscape","subcat":"Turfing","name":"Cow grass turfing S&I","method":"Cow Grass","unit":"sqft","supply":"supply_install","kl_min":1.5,"kl_max":3,"sg_min":None,"sg_max":None},
    {"cat":"Landscape","subcat":"Turfing","name":"Japanese grass turfing S&I","method":"Japanese Grass","unit":"sqft","supply":"supply_install","kl_min":2.5,"kl_max":4.5,"sg_min":None,"sg_max":None},
    {"cat":"Landscape","subcat":"Paving","name":"Granite/natural stone paving S&I","method":"Granite/Sandstone","unit":"sqft","supply":"supply_install","kl_min":15,"kl_max":28,"sg_min":None,"sg_max":None},
    {"cat":"Landscape","subcat":"Drainage","name":"Drainage channel U-channel","method":"U-channel + Grate","unit":"m","supply":"supply_install","kl_min":40,"kl_max":80,"sg_min":None,"sg_max":None},
    # Cleaning (6)
    {"cat":"Cleaning","subcat":"Post-Reno","name":"Post-renovation cleaning basic","method":"Dust + Wipe","unit":"sqft","supply":"supply_install","kl_min":0.6,"kl_max":1.3,"sg_min":0.25,"sg_max":0.50},
    {"cat":"Cleaning","subcat":"Post-Reno","name":"Post-renovation cleaning thorough","method":"Full Detail","unit":"sqft","supply":"supply_install","kl_min":1.2,"kl_max":2,"sg_min":0.40,"sg_max":0.80},
    {"cat":"Cleaning","subcat":"Floor","name":"Chemical wash tiles/floors","method":"Acid/Chemical","unit":"sqft","supply":"supply_install","kl_min":0.8,"kl_max":1.5,"sg_min":0.20,"sg_max":0.50},
    {"cat":"Cleaning","subcat":"Floor","name":"Marble floor polishing","method":"Diamond Polish","unit":"sqft","supply":"supply_install","kl_min":3,"kl_max":6,"sg_min":2.00,"sg_max":6.00},
    {"cat":"Cleaning","subcat":"Floor","name":"Parquet floor refinishing/varnishing","method":"Sand + Varnish","unit":"sqft","supply":"supply_install","kl_min":2.5,"kl_max":5,"sg_min":1.50,"sg_max":4.00},
    {"cat":"Cleaning","subcat":"Post-Reno","name":"Post-reno full unit cleaning","method":"Full Unit","unit":"lump sum","supply":"supply_install","kl_min":300,"kl_max":700,"sg_min":None,"sg_max":None},
]

# ---------------------------------------------------------------------------
# CATEGORY COLOURS
# ---------------------------------------------------------------------------
CAT_COLORS = {
    "Construction":    "E8F4FD",
    "Demolition":      "FFF3E0",
    "Tiling":          "F3E5F5",
    "Waterproofing":   "E8F5E9",
    "Electrical":      "FFF9C4",
    "Plumbing":        "E3F2FD",
    "Painting":        "FCE4EC",
    "False Ceiling":   "F5F5F5",
    "Carpentry":       "FBE9E7",
    "Table Top":       "F9FBE7",
    "Roofing":         "E0F2F1",
    "Aluminium":       "E8EAF6",
    "Glass":           "EDE7F6",
    "Flooring":        "FFF8E1",
    "Air Conditioning":"E1F5FE",
    "Metal Work":      "EFEBE9",
    "Landscape":       "E8F5E9",
    "Cleaning":        "F3E5F5",
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_side(left=False, right=False, top=False, bottom=False):
    thick = Side(style="medium")
    thin  = Side(style="thin")
    return Border(
        left   = thick if left   else thin,
        right  = thick if right  else thin,
        top    = thick if top    else thin,
        bottom = thick if bottom else thin,
    )

def avg_val(mn, mx):
    if mn is None or mx is None:
        return None
    return (mn + mx) / 2.0

def fmt_num(val):
    """Return float rounded to 2dp, or None."""
    if val is None:
        return None
    return round(val, 2)

# ---------------------------------------------------------------------------
# SHEET 1 — PRICE GUIDE
# ---------------------------------------------------------------------------
def build_price_guide(wb):
    ws = wb.create_sheet("Price Guide")

    # ---- column widths ----
    col_widths = {
        "A": 20, "B": 22, "C": 30, "D": 28, "E": 8, "F": 16,
        "G": 10, "H": 10, "I": 10,   # KL
        "J": 10, "K": 10, "L": 10,   # PG
        "M": 10, "N": 10, "O": 10,   # JB
        "P": 10, "Q": 10, "R": 10,   # SG
        "S": 8,  "T": 12,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ---- ROW 1: region group headers ----
    hdr1_fill = _fill("1F4E79")
    hdr1_font = _font(bold=True, size=13, color="FFFFFF")
    hdr1_align = _align("center", "center")

    static_headers = ["Category", "Subcategory", "Item Name", "Material/Method", "Unit", "Supply Type"]
    for ci, h in enumerate(static_headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr1_fill
        cell.font = hdr1_font
        cell.alignment = hdr1_align

    # Merged region headers
    region_groups = [
        (7, 9,  "KL / Selangor (RM)"),
        (10, 12, "Penang (RM)"),
        (13, 15, "JB / Johor (RM)"),
        (16, 18, "Singapore (SGD)"),
    ]
    for (start_col, end_col, label) in region_groups:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.fill = hdr1_fill
        cell.font = hdr1_font
        cell.alignment = hdr1_align

    for ci in [19, 20]:
        cell = ws.cell(row=1, column=ci, value="Samples" if ci == 19 else "Confidence")
        cell.fill = hdr1_fill
        cell.font = hdr1_font
        cell.alignment = hdr1_align

    # ---- ROW 2: sub-headers ----
    hdr2_fill = _fill("2E75B6")
    hdr2_font = _font(bold=False, size=10, color="FFFFFF")

    sub_labels = [
        "", "", "", "", "", "",
        "Min", "Max", "Avg",
        "Min", "Max", "Avg",
        "Min", "Max", "Avg",
        "Min", "Max", "Avg",
        "", "",
    ]
    for ci, label in enumerate(sub_labels, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill = hdr2_fill
        cell.font = hdr2_font
        cell.alignment = _align("center", "center")

    # ---- DATA ROWS ----
    NO_DATA_FILL   = _fill("F0F0F0")
    NO_DATA_FONT   = _font(italic=True, color="999999")
    NUM_FORMAT     = "0.00"
    HIGH_CONF_FILL = _fill("90EE90")
    data_start_row = 3

    # Track contiguous runs for merging (col A = category, col B = subcategory).
    # A new run starts whenever the value changes from the previous row.
    cat_runs    = []  # each entry: [r1, r2, cat]
    subcat_runs = []  # each entry: [r1, r2, cat, subcat]
    prev_cat = prev_subcat = prev_sc_cat = None

    # Pass 1: write data cells and build run metadata
    for idx, item in enumerate(ITEMS):
        row    = data_start_row + idx
        cat    = item["cat"]
        subcat = item["subcat"]

        # Maintain contiguous-run bookkeeping
        if cat != prev_cat:
            cat_runs.append([row, row, cat])
            prev_cat = cat
        else:
            cat_runs[-1][1] = row

        if subcat != prev_subcat or cat != prev_sc_cat:
            subcat_runs.append([row, row, cat, subcat])
            prev_subcat = subcat
            prev_sc_cat = cat
        else:
            subcat_runs[-1][1] = row

        kl_min = item["kl_min"]
        kl_max = item["kl_max"]
        kl_avg = avg_val(kl_min, kl_max)
        sg_min = item["sg_min"]
        sg_max = item["sg_max"]
        sg_avg = avg_val(sg_min, sg_max)

        row_fill = _fill(CAT_COLORS.get(cat, "FFFFFF"))

        def set_cell(col, value, num=False, no_data=False, _row=row, _rf=row_fill):
            c = ws.cell(row=_row, column=col, value=value)
            c.fill = _rf
            c.alignment = _align("right" if num else "left", "center")
            if no_data:
                c.fill = NO_DATA_FILL
                c.font = NO_DATA_FONT
                c.alignment = _align("center", "center")
            else:
                c.font = _font()
            if num and value is not None:
                c.number_format = NUM_FORMAT
            return c

        set_cell(1,  cat)
        set_cell(2,  subcat)
        set_cell(3,  item["name"])
        set_cell(4,  item["method"])
        set_cell(5,  item["unit"])
        set_cell(6,  item["supply"])
        set_cell(7,  fmt_num(kl_min), num=True)
        set_cell(8,  fmt_num(kl_max), num=True)
        set_cell(9,  fmt_num(kl_avg), num=True)
        set_cell(10, fmt_num(kl_min), num=True)  # PG = KL
        set_cell(11, fmt_num(kl_max), num=True)
        set_cell(12, fmt_num(kl_avg), num=True)
        set_cell(13, fmt_num(kl_min), num=True)  # JB = KL
        set_cell(14, fmt_num(kl_max), num=True)
        set_cell(15, fmt_num(kl_avg), num=True)

        if sg_min is None:
            set_cell(16, "no data", no_data=True)
            set_cell(17, "no data", no_data=True)
            set_cell(18, "no data", no_data=True)
        else:
            set_cell(16, fmt_num(sg_min), num=True)
            set_cell(17, fmt_num(sg_max), num=True)
            set_cell(18, fmt_num(sg_avg), num=True)

        s_cell = ws.cell(row=row, column=19, value=85)
        s_cell.fill = row_fill
        s_cell.font = _font()
        s_cell.alignment = _align("center", "center")

        conf_cell = ws.cell(row=row, column=20, value="high")
        conf_cell.fill = HIGH_CONF_FILL
        conf_cell.font = _font(bold=True)
        conf_cell.alignment = _align("center", "center")

    # Pass 2: merge contiguous runs in col A (category)
    for (r1, r2, cat) in cat_runs:
        if r1 < r2:
            ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        cell = ws.cell(row=r1, column=1)
        cell.value = cat
        cell.fill = _fill(CAT_COLORS.get(cat, "FFFFFF"))
        cell.font = _font(bold=True)
        cell.alignment = _align("center", "center", wrap=True)

    # Pass 3: merge contiguous runs in col B (subcategory)
    for (r1, r2, cat, subcat) in subcat_runs:
        if r1 < r2:
            ws.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=2)
        cell = ws.cell(row=r1, column=2)
        cell.value = subcat
        cell.fill = _fill(CAT_COLORS.get(cat, "FFFFFF"))
        cell.font = _font()
        cell.alignment = _align("center", "center", wrap=True)

    # ---- Freeze panes ----
    ws.freeze_panes = "A3"

    # ---- Row heights ----
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16


# ---------------------------------------------------------------------------
# SHEET 2 — DB SEED
# ---------------------------------------------------------------------------
def build_db_seed(wb):
    ws = wb.create_sheet("DB Seed")

    headers = [
        "Item Name", "Category", "Subcategory", "Material/Method",
        "Unit", "Supply Type", "Region",
        "Min Price", "Max Price", "Avg Price",
        "Samples", "Confidence"
    ]

    col_widths = [32, 18, 20, 28, 8, 16, 10, 10, 10, 10, 8, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row
    hdr_fill = _fill("1F4E79")
    hdr_font = _font(bold=True, size=11, color="FFFFFF")
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = _align("center", "center")

    ws.freeze_panes = "A2"

    row_num = 2
    LIGHT_GREY = "F5F5F5"

    for item in ITEMS:
        kl_min = item["kl_min"]
        kl_max = item["kl_max"]
        kl_avg = avg_val(kl_min, kl_max)
        sg_min = item["sg_min"]
        sg_max = item["sg_max"]
        sg_avg = avg_val(sg_min, sg_max)

        regions = [
            ("MY_KL", kl_min, kl_max, kl_avg),
            ("MY_PG", kl_min, kl_max, kl_avg),
            ("MY_JB", kl_min, kl_max, kl_avg),
            ("SG",   sg_min, sg_max, sg_avg),
        ]

        for region, mn, mx, av in regions:
            fill_color = LIGHT_GREY if (row_num % 2 == 0) else "FFFFFF"
            row_fill = _fill(fill_color)

            def sc(col, value, num=False, no_data=False):
                c = ws.cell(row=row_num, column=col, value=value)
                c.fill = row_fill
                c.alignment = _align("right" if num else "left", "center")
                if no_data:
                    c.font = _font(italic=True, color="999999")
                    c.alignment = _align("center", "center")
                else:
                    c.font = _font()
                if num and value is not None and not no_data:
                    c.number_format = "0.00"
                return c

            sc(1,  item["name"])
            sc(2,  item["cat"])
            sc(3,  item["subcat"])
            sc(4,  item["method"])
            sc(5,  item["unit"])
            sc(6,  item["supply"])
            sc(7,  region)

            if mn is None:
                sc(8,  "no data", no_data=True)
                sc(9,  "no data", no_data=True)
                sc(10, "no data", no_data=True)
            else:
                sc(8,  fmt_num(mn), num=True)
                sc(9,  fmt_num(mx), num=True)
                sc(10, fmt_num(av), num=True)

            sc(11, 85, num=True)
            conf_c = ws.cell(row=row_num, column=12, value="high")
            conf_c.fill = _fill("90EE90")
            conf_c.font = _font(bold=True)
            conf_c.alignment = _align("center", "center")

            row_num += 1

    ws.row_dimensions[1].height = 18


# ---------------------------------------------------------------------------
# SHEET 3 — SUMMARY
# ---------------------------------------------------------------------------
def build_summary(wb):
    ws = wb.create_sheet("Summary")

    headers = ["Category", "Item Count", "KL Price Range", "SG Coverage (%)", "Notes"]
    col_widths = [22, 12, 22, 18, 45]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header
    hdr_fill = _fill("1F4E79")
    hdr_font = _font(bold=True, size=11, color="FFFFFF")
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = _align("center", "center")

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # Build per-category stats
    from collections import defaultdict
    cat_items = defaultdict(list)
    for item in ITEMS:
        cat_items[item["cat"]].append(item)

    NOTES = {
        "Construction":    "Structural works; prices vary by design complexity",
        "Demolition":      "Labour-only trades; disposal charged separately",
        "Tiling":          "Wide range due to tile grade & format size",
        "Waterproofing":   "Critical pre-tile work; warranty essential",
        "Electrical":      "DB upgrade may be required for older units",
        "Plumbing":        "SG prices per unit significantly higher than MY",
        "Painting":        "Skim coat adds cost; always confirm coat count",
        "False Ceiling":   "SG prices quoted per linear ft for cove design",
        "Carpentry":       "Largest cost item; plywood grade affects price",
        "Table Top":       "Natural stone priced per sqft; quartz by lin ft",
        "Roofing":         "MY-specific items common for landed properties",
        "Aluminium":       "SG prices per unit not per sqft; verify with supplier",
        "Glass":           "Thickness affects price; 10mm ≈ 25% more than 8mm",
        "Flooring":        "SPC most popular; solid timber premium product",
        "Air Conditioning":"System pricing; SG data incomplete for multi-split",
        "Metal Work":      "Powder coat standard; SS commands premium",
        "Landscape":       "SG data limited; prices vary by garden size",
        "Cleaning":        "SG post-reno rates much higher per sqft than MY",
    }

    row_num = 2
    ALT_FILL = "F5F5F5"

    for cat, items in cat_items.items():
        kl_mins = [i["kl_min"] for i in items]
        kl_maxs = [i["kl_max"] for i in items]
        kl_range = f"RM {min(kl_mins):.2f} – {max(kl_maxs):.2f}"

        sg_count = sum(1 for i in items if i["sg_min"] is not None)
        sg_pct = round(sg_count / len(items) * 100)

        fill_color = ALT_FILL if (row_num % 2 == 0) else "FFFFFF"
        row_fill = _fill(fill_color)
        cat_fill = _fill(CAT_COLORS.get(cat, "FFFFFF"))

        c1 = ws.cell(row=row_num, column=1, value=cat)
        c1.fill = cat_fill
        c1.font = _font(bold=True)
        c1.alignment = _align("left", "center")

        c2 = ws.cell(row=row_num, column=2, value=len(items))
        c2.fill = row_fill
        c2.font = _font()
        c2.alignment = _align("center", "center")

        c3 = ws.cell(row=row_num, column=3, value=kl_range)
        c3.fill = row_fill
        c3.font = _font()
        c3.alignment = _align("left", "center")

        # SG coverage badge
        if sg_pct >= 70:
            sg_fill = _fill("90EE90")
        elif sg_pct >= 30:
            sg_fill = _fill("FFD700")
        else:
            sg_fill = _fill("FFB6C1")

        c4 = ws.cell(row=row_num, column=4, value=f"{sg_pct}%")
        c4.fill = sg_fill
        c4.font = _font(bold=True)
        c4.alignment = _align("center", "center")

        c5 = ws.cell(row=row_num, column=5, value=NOTES.get(cat, ""))
        c5.fill = row_fill
        c5.font = _font()
        c5.alignment = _align("left", "center", wrap=True)

        ws.row_dimensions[row_num].height = 18
        row_num += 1


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "docs", "price_database_seed.xlsx"
    )

    wb = Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    print("Building Sheet 1: Price Guide ...")
    build_price_guide(wb)

    print("Building Sheet 2: DB Seed ...")
    build_db_seed(wb)

    print("Building Sheet 3: Summary ...")
    build_summary(wb)

    print(f"Saving to {out_path} ...")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

    # Quick verification
    from openpyxl import load_workbook
    verify = load_workbook(out_path)
    sheets = verify.sheetnames
    pg   = verify["Price Guide"]
    seed = verify["DB Seed"]
    summ = verify["Summary"]

    data_rows_pg   = pg.max_row   - 2
    data_rows_seed = seed.max_row - 1
    data_rows_summ = summ.max_row - 1

    print(f"\nVerification:")
    print(f"  Sheets       : {sheets}")
    print(f"  Price Guide  : {data_rows_pg} data rows  (expected {len(ITEMS)})")
    print(f"  DB Seed      : {data_rows_seed} data rows  (expected {len(ITEMS) * 4})")
    print(f"  Summary      : {data_rows_summ} category rows")
    print(f"  File size    : {os.path.getsize(out_path):,} bytes")
    print(f"\nDone -> {out_path}")

    # Count unique categories in summary
    cats_in_summ = set()
    for row in summ.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cats_in_summ.add(row[0])
    print(f"  Categories in Summary: {sorted(cats_in_summ)}")

    assert data_rows_pg   == len(ITEMS),       f"Row count mismatch in Price Guide"
    assert data_rows_seed == len(ITEMS) * 4,   f"Row count mismatch in DB Seed"
    print("\nAll assertions passed.")


if __name__ == "__main__":
    main()
