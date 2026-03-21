from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Price Reference"

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="0F1923")
data_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 14

# Title
ws.merge_cells("A1:H1")
title = ws["A1"]
title.value = "RenoSmart PRICE_REFERENCE - Designer Quotation Prices (2025-2026)"
title.font = Font(name="Arial", bold=True, size=13, color="F0B90B")
title.fill = PatternFill("solid", fgColor="0F1923")
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

headers = ["Category", "Item", "Supply Type", "Unit", "MY Min (RM)", "MY Max (RM)", "SG Min ($)", "SG Max ($)"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

ws.freeze_panes = "A3"

data = [
    ("Tiling", "Tiling LABOUR ONLY (floor 300x300-600x600)", "Labour Only", "sqft", 6, 10, 8, 15),
    ("Tiling", "Tiling LABOUR ONLY (wall 300x600-600x600)", "Labour Only", "sqft", 7, 12, 10, 18),
    ("Tiling", "Tiling LABOUR ONLY (large format 900x900+)", "Labour Only", "sqft", 8, 14, 12, 20),
    ("Tiling", "Tiling S&I (floor 300x300)", "S&I", "sqft", 15, 22, 20, 32),
    ("Tiling", "Tiling S&I (floor 300x600)", "S&I", "sqft", 16, 24, 22, 35),
    ("Tiling", "Tiling S&I (floor 600x600)", "S&I", "sqft", 18, 28, 25, 40),
    ("Tiling", "Tiling S&I (floor 800x800)", "S&I", "sqft", 20, 30, 28, 42),
    ("Tiling", "Tiling S&I (floor 900x900)", "S&I", "sqft", 22, 32, 30, 45),
    ("Tiling", "Tiling S&I (floor 1200x600)", "S&I", "sqft", 25, 35, 32, 48),
    ("Tiling", "Tiling S&I (wall 200x300)", "S&I", "sqft", 15, 22, 20, 32),
    ("Tiling", "Tiling S&I (wall 300x600)", "S&I", "sqft", 18, 26, 24, 38),
    ("Tiling", "Tiling S&I (wall 600x600)", "S&I", "sqft", 20, 28, 28, 40),
    ("Tiling", "Tiling S&I (wall 600x1200)", "S&I", "sqft", 25, 35, 32, 48),
    ("Tiling", "Tiling S&I (outdoor anti-slip)", "S&I", "sqft", 18, 30, 25, 42),
    ("Tiling", "Tiling S&I (feature/mosaic)", "S&I", "sqft", 25, 45, 35, 60),
    ("Tiling", "Tiling S&I (rubber tiles)", "S&I", "sqft", 25, 40, 35, 55),
    ("Construction", "Build-up brick wall c/w plastering", "S&I", "sqft", 20, 35, 30, 50),
    ("Construction", "Build-up brick wall (labour only)", "Labour Only", "sqft", 12, 20, 18, 30),
    ("Construction", "RC floor slab 150mm G25", "S&I", "sqft", 35, 55, 50, 75),
    ("Construction", "RC floor slab 150mm G30", "S&I", "sqft", 38, 60, 55, 80),
    ("Construction", "RC slab extension", "S&I", "sqft", 40, 65, 55, 85),
    ("Construction", "Concrete flooring c/w BRC", "S&I", "sqft", 15, 35, 22, 48),
    ("Construction", "Screed/Re-leveling", "S&I", "sqft", 8, 15, 12, 22),
    ("Construction", "Plastering finish", "S&I", "sqft", 5, 10, 8, 15),
    ("Construction", "Kerb (brick/concrete)", "S&I", "ft", 40, 80, 55, 100),
    ("Electrical", "Lighting point S&I (new wiring + cable)", "S&I", "pt", 55, 100, 80, 150),
    ("Electrical", "Lighting point S&I (>12ft height)", "S&I", "pt", 80, 130, 110, 180),
    ("Electrical", "13A socket power point S&I", "S&I", "pt", 85, 150, 130, 220),
    ("Electrical", "15A socket power point S&I", "S&I", "pt", 100, 170, 150, 250),
    ("Electrical", "USB socket point S&I", "S&I", "pt", 120, 200, 180, 280),
    ("Electrical", "Ceiling fan point S&I", "S&I", "pt", 100, 200, 150, 280),
    ("Electrical", "Exhaust fan point S&I", "S&I", "pt", 80, 150, 120, 220),
    ("Electrical", "Cove light point S&I", "S&I", "pt", 120, 250, 180, 350),
    ("Electrical", "Downlight cutout + wiring", "S&I", "pt", 35, 60, 50, 85),
    ("Electrical", "Re-locate existing point", "S&I", "pt", 150, 300, 200, 400),
    ("Electrical", "Data/internet point (Cat5e/6)", "S&I", "pt", 200, 400, 300, 600),
    ("Electrical", "TV point S&I", "S&I", "pt", 150, 300, 200, 450),
    ("Electrical", "Electrical labour only (install fitting)", "Labour Only", "pt", 15, 35, 25, 50),
    ("Electrical", "DB box upgrade (18-way)", "S&I", "unit", 800, 1500, 1200, 2000),
    ("Electrical", "DB box upgrade (36-way)", "S&I", "unit", 1200, 2200, 1800, 3000),
    ("Painting", "Painting (interior 2-coat)", "S&I", "sqft", 2.5, 5, 3.5, 7),
    ("Painting", "Painting (skim coat + paint)", "S&I", "sqft", 4, 8, 6, 12),
    ("Painting", "Painting (exterior weather-shield)", "S&I", "sqft", 3, 6, 5, 9),
    ("Painting", "Re-paint existing roofing", "S&I", "sqft", 2, 4, 3, 6),
    ("False Ceiling", "False ceiling (flat plasterboard)", "S&I", "sqft", 10, 18, 15, 25),
    ("False Ceiling", "False ceiling c/w cove light", "S&I", "sqft", 14, 22, 20, 32),
    ("False Ceiling", "False ceiling (L-box/design)", "S&I", "sqft", 18, 35, 25, 50),
    ("False Ceiling", "False ceiling (calcium silicate)", "S&I", "sqft", 12, 20, 18, 28),
    ("False Ceiling", "Partition wall (single layer)", "S&I", "sqft", 18, 30, 25, 42),
    ("Carpentry", "Kitchen cabinet (laminated/melamine)", "S&I", "ft", 450, 950, 700, 1400),
    ("Carpentry", "Kitchen cabinet (solid plywood + veneer)", "S&I", "ft", 550, 1200, 800, 1800),
    ("Carpentry", "Kitchen cabinet (aluminium honeycomb/ACP)", "S&I", "ft", 400, 900, 600, 1300),
    ("Carpentry", "Kitchen cabinet (solid nyatoh/rubber wood)", "S&I", "ft", 800, 1800, 1200, 2500),
    ("Carpentry", "Wardrobe (laminated swing door)", "S&I", "ft", 700, 1200, 1000, 1800),
    ("Carpentry", "Wardrobe (laminated sliding)", "S&I", "ft", 800, 1400, 1200, 2000),
    ("Carpentry", "Wardrobe (solid plywood)", "S&I", "ft", 900, 1600, 1300, 2300),
    ("Carpentry", "TV console/feature wall", "S&I", "ft", 200, 500, 300, 750),
    ("Carpentry", "Shoe cabinet", "S&I", "ft", 250, 550, 400, 800),
    ("Carpentry", "Vanity cabinet", "S&I", "ft", 350, 700, 500, 1000),
    ("Plumbing", "Basin + mixer tap install", "S&I", "unit", 300, 600, 500, 900),
    ("Plumbing", "WC floor-mount", "S&I", "unit", 400, 800, 600, 1200),
    ("Plumbing", "WC wall-hung", "S&I", "unit", 800, 1500, 1200, 2200),
    ("Plumbing", "Water heater install", "S&I", "unit", 200, 500, 350, 700),
    ("Plumbing", "Rain shower set", "S&I", "unit", 400, 900, 600, 1300),
    ("Plumbing", "Toilet bowl outlet piping", "S&I", "L-sum", 800, 2000, 1200, 3000),
    ("Plumbing", "Inlet/outlet pipe for basin/tap", "S&I", "L-sum", 400, 1000, 600, 1500),
    ("Waterproofing", "Bathroom (cementitious)", "S&I", "sqft", 8, 18, 12, 25),
    ("Waterproofing", "Flat roof (torch-on membrane)", "S&I", "sqft", 12, 25, 18, 35),
    ("Waterproofing", "Balcony/patio", "S&I", "sqft", 10, 20, 15, 28),
    ("Demolition", "Hacking floor", "S&I", "sqft", 3, 8, 5, 12),
    ("Demolition", "Hacking wall", "S&I", "sqft", 5, 12, 8, 18),
    ("Demolition", "Hacking + re-leveling", "S&I", "sqft", 8, 15, 12, 22),
    ("Roofing", "Metal roofing (PU metal deck)", "S&I", "sqft", 22, 34, 30, 48),
    ("Roofing", "Polycarbonate roof (twinwall)", "S&I", "sqft", 18, 30, 25, 42),
    ("Roofing", "Glass roofing (tempered)", "S&I", "sqft", 55, 90, 75, 125),
    ("Aluminium & Glass", "Aluminium casement window", "S&I", "sqft", 100, 250, 150, 350),
    ("Aluminium & Glass", "Aluminium sliding door", "S&I", "sqft", 120, 300, 180, 400),
    ("Aluminium & Glass", "Glass shower screen (10mm)", "S&I", "sqft", 45, 90, 65, 130),
    ("Aluminium & Glass", "Tempered glass fix panel", "S&I", "sqft", 55, 100, 75, 140),
    ("Aluminium & Glass", "Mirror (standard)", "S&I", "sqft", 22, 40, 30, 55),
    ("Flooring", "Vinyl SPC click-lock", "S&I", "sqft", 6, 12, 8, 18),
    ("Flooring", "Timber engineered", "S&I", "sqft", 12, 28, 18, 40),
    ("Flooring", "Carpet grass (artificial turf)", "S&I", "sqft", 8, 18, 12, 25),
    ("Flooring", "Artificial grass", "S&I", "sqft", 15, 35, 22, 48),
    ("Flooring", "Composite decking", "S&I", "sqft", 45, 90, 60, 120),
    ("Others", "Air conditioning 1.5HP split inverter", "S&I", "unit", 1800, 3500, 2500, 4500),
    ("Others", "Metal railing mild steel powder coated", "S&I", "ft", 80, 180, 120, 250),
    ("Others", "Stainless steel fencing", "S&I", "sqft", 35, 80, 50, 110),
    ("Others", "PVC wall panel", "S&I", "sqft", 20, 50, 30, 70),
    ("Others", "Staircase (metal I-beam structure)", "S&I", "L-sum", 15000, 35000, 22000, 50000),
    ("Others", "Staircase (wood step + handrail)", "S&I", "L-sum", 30000, 60000, 45000, 85000),
    ("Others", "Landscape soil + carpet grass", "S&I", "sqft", 8, 18, 12, 25),
]

cat_colors = {
    "Tiling": "E8F5E9", "Construction": "E3F2FD", "Electrical": "FFF3E0",
    "Painting": "F3E5F5", "False Ceiling": "E0F7FA", "Carpentry": "FBE9E7",
    "Plumbing": "E8EAF6", "Waterproofing": "E0F2F1", "Demolition": "FFEBEE",
    "Roofing": "F1F8E9", "Aluminium & Glass": "E1F5FE", "Flooring": "FFF8E1",
    "Others": "F5F5F5",
}

for i, row in enumerate(data, 3):
    cat, item, stype, unit, my_min, my_max, sg_min, sg_max = row
    fill = PatternFill("solid", fgColor=cat_colors.get(cat, "FFFFFF"))
    for col in range(1, 9):
        ws.cell(row=i, column=col).border = border
        ws.cell(row=i, column=col).fill = fill

    ws.cell(row=i, column=1, value=cat).font = bold_font
    ws.cell(row=i, column=1).alignment = left_align
    ws.cell(row=i, column=2, value=item).font = data_font
    ws.cell(row=i, column=2).alignment = left_align
    ws.cell(row=i, column=3, value=stype).font = data_font
    ws.cell(row=i, column=3).alignment = center
    ws.cell(row=i, column=4, value=unit).font = data_font
    ws.cell(row=i, column=4).alignment = center

    for col, val in [(5, my_min), (6, my_max), (7, sg_min), (8, sg_max)]:
        c = ws.cell(row=i, column=col, value=val)
        c.font = data_font
        c.alignment = center
        c.number_format = "#,##0.00" if isinstance(val, float) else "#,##0"

ws.auto_filter.ref = f"A2:H{len(data) + 2}"

wb.save("docs/PRICE_REFERENCE_2025.xlsx")
print("Exported to docs/PRICE_REFERENCE_2025.xlsx")
