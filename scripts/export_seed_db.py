import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Parse seed SQL
rows = []
with open("supabase/seed_price_database.sql", "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line.startswith("('"): continue
        # Remove trailing comma or semicolon
        line = line.rstrip(",;")
        # Parse: ('item_name','category','subcategory','material_method','unit','supply_type','region',min,max,avg,samples,'confidence',NOW())
        m = re.match(
            r"\('(.+?)','(.+?)','(.+?)','(.+?)','(.+?)','(.+?)','(.+?)',"
            r"([\d.]+),([\d.]+),([\d.]+),(\d+),'(\w+)',NOW\(\)\)",
            line
        )
        if m:
            rows.append({
                "item_name": m.group(1),
                "category": m.group(2),
                "subcategory": m.group(3),
                "material_method": m.group(4),
                "unit": m.group(5),
                "supply_type": m.group(6),
                "region": m.group(7),
                "min_price": float(m.group(8)),
                "max_price": float(m.group(9)),
                "avg_price": float(m.group(10)),
                "sample_count": int(m.group(11)),
                "confidence": m.group(12),
            })

print(f"Parsed {len(rows)} rows")

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "price_database"

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="0F1923")
data_font = Font(name="Arial", size=9)
bold_font = Font(name="Arial", bold=True, size=9)
border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
center = Alignment(horizontal="center", vertical="center")
left_a = Alignment(horizontal="left", vertical="center")

# Title
ws.merge_cells("A1:L1")
title = ws["A1"]
title.value = "RenoSmart price_database (Seed Data) - All Regions"
title.font = Font(name="Arial", bold=True, size=13, color="F0B90B")
title.fill = PatternFill("solid", fgColor="0F1923")
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

headers = ["Item Name", "Category", "Subcategory", "Material/Method", "Unit", "Supply Type", "Region", "Min Price", "Max Price", "Avg Price", "Samples", "Confidence"]
widths = [38, 16, 20, 22, 8, 14, 8, 12, 12, 12, 10, 12]
for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[chr(64 + col) if col <= 26 else None].width = w

# Fix column L width
from openpyxl.utils import get_column_letter
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A3"

cat_colors = {
    "Construction": "E3F2FD", "Tiling": "E8F5E9", "Electrical": "FFF3E0",
    "Painting": "F3E5F5", "False Ceiling": "E0F7FA", "Carpentry": "FBE9E7",
    "Plumbing": "E8EAF6", "Waterproofing": "E0F2F1", "Demolition": "FFEBEE",
    "Roofing": "F1F8E9", "Aluminium": "E1F5FE", "Flooring": "FFF8E1",
    "Glass": "E1F5FE", "Metal Work": "EFEBE9", "Door": "FFF8E1",
    "Air Conditioning": "F3E5F5", "Stone": "EFEBE9", "Landscape": "E8F5E9",
    "Alarm & CCTV": "FFF3E0", "Cleaning": "F5F5F5",
}

conf_colors = {"high": "4CAF50", "mid": "FF9800", "low": "F44336"}

for i, row in enumerate(rows, 3):
    fill_color = "FFFFFF"
    for key, color in cat_colors.items():
        if key in row["category"]:
            fill_color = color
            break
    fill = PatternFill("solid", fgColor=fill_color)

    vals = [
        row["item_name"], row["category"], row["subcategory"], row["material_method"],
        row["unit"], row["supply_type"], row["region"],
        row["min_price"], row["max_price"], row["avg_price"],
        row["sample_count"], row["confidence"],
    ]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.font = bold_font if col <= 2 else data_font
        c.fill = fill
        c.border = border
        c.alignment = left_a if col <= 4 else center
        if col in (8, 9, 10):
            c.number_format = "#,##0.00"
        if col == 12:
            conf_c = conf_colors.get(val, "000000")
            c.font = Font(name="Arial", bold=True, size=9, color=conf_c)

ws.auto_filter.ref = f"A2:L{len(rows) + 2}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.merge_cells("A1:D1")
ws2["A1"].value = "Price Database Summary"
ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="F0B90B")
ws2["A1"].fill = PatternFill("solid", fgColor="0F1923")

# Count by category
from collections import Counter
cat_counts = Counter(r["category"] for r in rows)
region_counts = Counter(r["region"] for r in rows)

ws2["A3"] = "Category"
ws2["B3"] = "Count"
ws2["A3"].font = header_font
ws2["B3"].font = header_font
ws2["A3"].fill = header_fill
ws2["B3"].fill = header_fill
for i, (cat, cnt) in enumerate(sorted(cat_counts.items()), 4):
    ws2.cell(row=i, column=1, value=cat).font = data_font
    ws2.cell(row=i, column=2, value=cnt).font = data_font

r = i + 2
ws2.cell(row=r, column=1, value="Region").font = header_font
ws2.cell(row=r, column=2, value="Count").font = header_font
ws2.cell(row=r, column=1).fill = header_fill
ws2.cell(row=r, column=2).fill = header_fill
for j, (reg, cnt) in enumerate(sorted(region_counts.items()), r + 1):
    ws2.cell(row=j, column=1, value=reg).font = data_font
    ws2.cell(row=j, column=2, value=cnt).font = data_font

ws2.cell(row=j + 2, column=1, value="Total Entries").font = bold_font
ws2.cell(row=j + 2, column=2, value=len(rows)).font = bold_font

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 12

wb.save("docs/price_database_seed.xlsx")
print(f"Exported {len(rows)} rows to docs/price_database_seed.xlsx")
